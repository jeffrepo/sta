# -*- encoding: utf-8 -*-

from openerp import models, fields, api, _
from odoo.exceptions import ValidationError, UserError, Warning
from openerp.osv.orm import except_orm
import time
import base64
import xlwt
import io
import logging
import requests
import json
from datetime import date
import datetime
import datetime
from datetime import datetime
import glob
import time
import openpyxl

import base64
import xlsxwriter
import io
import logging
import xlrd
import base64

class StaSyncImportacion(models.TransientModel):
    _name = 'sta.sync.importacion'

    archivo = fields.Binary('Archivo excel')

    @api.multi
    def codigo_productos(self):
        productos = self.env['product.template'].search([('default_code','=',False)])
        if productos:
            logging.warn(len(productos))
            for p in productos:
                strin_split = p.name.split()
                logging.warn(strin_split)
                if strin_split[0].isnumeric():
                    logging.warn(strin_split[0])
                    p.default_code = strin_split[0]

        return {
            'view_type': 'form',
            'view_mode': 'form',
            'res_model': 'sta.sync.importacion',
            'res_id': self.id,
            'view_id': False,
            'type': 'ir.actions.act_window',
            'target': 'new',
        }

    @api.multi
    def leer_excel_productos(self):
        workbook = xlrd.open_workbook(file_contents = base64.decodestring(self.archivo))
        sheet = workbook.sheet_by_index(0)
        productos = self.env['product.template'].search([('default_code','=',False)])
        if productos:
            logging.warn(len(productos))
            for p in productos:

                for linea in range(sheet.nrows):
                    if linea != 0:
                        # fecha_excel = sheet.cell(linea, 10).value
                        # fecha_excel = fecha_excel.replace("'", "")

                        # fecha = datetime.datetime(*xlrd.xldate_as_tuple(fecha_excel, workbook.datemode))
                        nombre_producto_excel = sheet.cell(linea, 1).value
                        codigo_producto_excel = sheet.cell(linea, 0).value

                        if str(p.name) == str(nombre_producto_excel):
                            logging.warn(p.name)
                            p.default_code = str(codigo_producto_excel)

        return {
            'view_type': 'form',
            'view_mode': 'form',
            'res_model': 'sta.sync.importacion',
            'res_id': self.id,
            'view_id': False,
            'type': 'ir.actions.act_window',
            'target': 'new',
        }

    @api.multi
    def crear_productos_existentes(self):
        workbook = xlrd.open_workbook(file_contents = base64.decodestring(self.archivo))
        sheet = workbook.sheet_by_index(0)
        productos = self.env['product.template'].search([('default_code','!=',False)])
        if productos:
            for p in productos:
                # logging.warn(p.default_code)
                # logging.warn(int(p.default_code))
                existe = False
                existe_producto = ''
                for linea in range(sheet.nrows):
                    if linea != 0:
                        nombre_producto_excel = sheet.cell(linea, 1).value
                        codigo_producto_excel = sheet.cell(linea, 0).value
                        existe_producto = codigo_producto_excel
                        logging.warn(existe_producto)
                        if int(p.default_code) == codigo_producto_excel:
                            existe = True



                # logging.warn(existe_producto)
                # logging.warn('no')

    def sync_importacion(self):
        path = "/opt/cuentas_bancarias_prestamos_comi.xlsx"
        wb_obj = openpyxl.load_workbook(path)
        sheet_obj = wb_obj.active
        cell_obj = sheet_obj.cell(row = 1, column = 1)
        logging.warn(cell_obj.value)
        numero_filas = sheet_obj.max_row
        numero_columnas = sheet_obj.max_column

        compra = {
        'partner_id': sheet_obj.cell(row = i, column = 6).value,
        'currency_id': sheet_obj.cell(row = i, column = 6).value,
        'picking_type_id': sheet_obj.cell(row = i, column = 6).value,
        }
        compra_id = self.env['purchase.order'].create(compra)
        for i in range(2,numero_filas+1):
            producto_id = False
            producto = self.env['product.product'].search([('name','=',nombre)])
            if len(product) > 0:
                product_id = producto
            else:
                product_id = self.env['product.product'].create({'name': nombre})
            linea = {
            'order_id': compra_id.id,
            'product_id': product_id.id,
            'product_qty': 0,
            'price_unit': 0
            }
            order_line = self.env['purchase.order.line'].create(linea)
    #     for i in range(2, numero_filas + 1):
    #         cell_obj = sheet_obj.cell(row = i, column = 1)
    #         id_sistema = sheet_obj.cell(row = i, column = 3).value
    #         logging.warn(id_sistema)
    #         numero_cuenta = sheet_obj.cell(row = i, column = 6).value
    #         existe_empleado = self.env['hr.employee'].search([('name','=',id_sistema)])
    #         valor_prestamo = sheet_obj.cell(row = i, column = 15).value
    #         if existe_empleado:
    #             logging.warn('si')
    #             if numero_cuenta:
    #                 partner_id = self.env['res.partner'].create({'name': existe_empleado.name})
    #                 cuenta = self.env['res.partner.bank'].create({'acc_number': numero_cuenta, 'acc_holder_name': existe_empleado.name,'partner_id': partner_id.id})
    #                 existe_empleado.write({'bank_account_id': cuenta.id })
    #                 cuenta.write({'company_id': existe_empleado.company_id.id})
    #                 logging.warn(cuenta)
    #     return True
    #
    # def sync_empleado(self):
    #     path = "/opt/emp.xlsx"
    #     wb_obj = openpyxl.load_workbook(path)
    #     sheet_obj = wb_obj.active
    #     cell_obj = sheet_obj.cell(row = 1, column = 1)
    #     logging.warn(cell_obj.value)
    #     numero_filas = sheet_obj.max_row
    #     numero_columnas = sheet_obj.max_column
    #
    #     for i in range(2, numero_filas + 1):
    #         cell_obj = sheet_obj.cell(row = i, column = 1)
    #         company = False
    #         struct_id = False
    #         if sheet_obj.cell(row = i, column = 2).value  == 'COANRO':
    #             company = 1
    #             struct_id = 2
    #         elif sheet_obj.cell(row = i, column = 2).value  == 'JONA':
    #             company = 4
    #             struct_id = 5
    #         elif sheet_obj.cell(row = i, column = 2).value  == 'LSM' or sheet_obj.cell(row = i, column = 2).value == 'lsm':
    #             company = 2
    #             struct_id = 4
    #         elif sheet_obj.cell(row = i, column = 2).value  == 'ADCOSINE':
    #             company = 3
    #             struct_id = 3
    #         elif sheet_obj.cell(row = i, column = 2).value  == 'LSM- VISTALAGO':
    #             company = 2
    #             struct_id = 4
    #         logging.warn(sheet_obj.cell(row = i, column = 4))
    #         empleado = {
    #             'name': str(sheet_obj.cell(row = i, column = 4).value) +' ' +str(sheet_obj.cell(row = i, column = 5).value)+' ' + str(sheet_obj.cell(row = i, column = 6).value) +' ' +  str(sheet_obj.cell(row = i, column = 7).value),
    #             'company_id': company,
    #             'id_sistema': sheet_obj.cell(row = i, column = 3).value,
    #             'primer_nombre': sheet_obj.cell(row = i, column = 4).value,
    #             'segundo_nombre': sheet_obj.cell(row = i, column = 5).value,
    #             'primer_apellido': sheet_obj.cell(row = i, column = 6).value,
    #             'segundo_apellido': sheet_obj.cell(row = i, column = 7).value,
    #             'apellido_casada': '',
    #             'country_id': 90,
    #             'estado_civil_igss': sheet_obj.cell(row = i, column = 9).value,
    #             'codigo_documento': sheet_obj.cell(row = i, column = 10).value,
    #             'identification_id': sheet_obj.cell(row = i, column = 11).value,
    #             'codigo_pais': sheet_obj.cell(row = i, column = 12).value,
    #             'codigo_nacimiento': sheet_obj.cell(row = i, column = 13).value,
    #             'nit': sheet_obj.cell(row = i, column = 14).value,
    #             'igss': sheet_obj.cell(row = i, column = 15).value,
    #             'gender': 'male' if sheet_obj.cell(row = i, column = 16).value == 'M' else 'female',
    #             'birthday': sheet_obj.cell(row = i, column = 17).value,
    #             'cantidad_hijos':  sheet_obj.cell(row = i, column = 18).value,
    #             'trabajo_extranjero': sheet_obj.cell(row = i, column = 19).value,
    #             'forma': '',
    #             'pais_trabajo': '',
    #             'motivo_retiro_extranjero': '',
    #             'etnia': sheet_obj.cell(row = i, column = 25).value,
    #             'idioma': sheet_obj.cell(row = i, column = 26).value,
    #             'puesto_codigo': sheet_obj.cell(row = i, column = 31).value
    #         }
    #         logging.warn(empleado)
    #         empleado_id = self.env['hr.employee'].create(empleado)
    #         logging.warn(empleado_id)
    #         contrato = {
    #             'name': 'Contrato' + ' '+str(empleado_id.name),
    #             'employee_id': empleado_id.id,
    #             'company_id': empleado_id.company_id.id,
    #             'date_start': sheet_obj.cell(row = i, column = 28).value,
    #             'tipo_contrato': sheet_obj.cell(row = i, column = 27).value,
    #             'date_end': sheet_obj.cell(row = i, column = 30).value,
    #             'wage': sheet_obj.cell(row = i, column = 41).value,
    #             'bonificacion_incentivo': 250,
    #             'struct_id': struct_id,
    #             'hora_extra': sheet_obj.cell(row = i, column = 43).value,
    #             'bono_otro': sheet_obj.cell(row = i, column = 39).value,
    #             'state': 'open'
    #         }
    #         contrato_id = self.env['hr.contract'].create(contrato)
    #         logging.warn(contrato_id.state)

        return {
            'context': self.env.context,
            'view_type': 'form',
            'view_mode': 'form',
            'res_model': 'sta.sync.importacion',
            'res_id': self.id,
            'view_id': False,
            'type': 'ir.actions.act_window',
            'target': 'new',
        }
